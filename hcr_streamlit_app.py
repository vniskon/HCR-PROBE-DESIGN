"""
HCR Split-Initiator Probe Designer — Streamlit App
Three input modes:
  1. Single FASTA  → auto-tiles, filters, ranks and designs probes
  2. Direct 52 bp target regions → designs probes straight from your targets
  3. Multi-FASTA   → finds common 52 bp regions across all sequences, then designs probes
"""

import io
import re
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.utils import get_column_letter

# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(page_title="HCR Probe Designer", page_icon="🧬", layout="wide")
st.title("🧬 HCR Split-Initiator Probe Designer")
st.markdown(
    "Three modes: design probes from a **single FASTA**, paste **direct 52 bp targets**, "
    "or upload a **multi-FASTA** to find common regions across all sequences first."
)

# ── HCR initiator sequences ────────────────────────────────────────────────────
INITIATORS = {
    "B1": {"P1": "GAGGAGGGCAGCAAACGG", "P2": "GAAGAGTCTTCCTTTACG", "S1": "AA", "S2": "TA"},
    "B2": {"P1": "CCTCGTAAATCCTCATCA", "P2": "ATCATCCAGTAAACCGCC", "S1": "AA", "S2": "AA"},
    "B3": {"P1": "GTCCCTGCCTCTATATCT", "P2": "CCACTCAACTTTAACCCG", "S1": "TT", "S2": "TT"},
    "B4": {"P1": "CCTCAACCTACCTCCAAC", "P2": "TCTCACCATATTCGCTTC", "S1": "AA", "S2": "AT"},
    "B5": {"P1": "CTCACTCCCAATCTCTAT", "P2": "CTACCCTACAAATCCAAT", "S1": "AA", "S2": "AA"},
}

# ── Colour palette ─────────────────────────────────────────────────────────────
C_TEAL      = "1A7A8A"
C_TEAL_LITE = "D0EDF1"
C_GREEN     = "1E8449"
C_GREEN_LT  = "D5F5E3"
C_PURPLE    = "6C3483"
C_PURPLE_LT = "E8DAEF"
C_GREY      = "F2F4F4"
C_WHITE     = "FFFFFF"
C_DARK      = "1C2833"
C_SPACER    = "E74C3C"
C_RED       = "C0392B"
C_RED_LT    = "FADBD8"

# ── Sequence utilities ─────────────────────────────────────────────────────────
_COMP = str.maketrans("ATGCatgc", "TACGtacg")

def parse_fasta(text):
    records = {}
    header, parts = None, []
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        if line.startswith(">"):
            if header is not None:
                records[header] = "".join(parts)
            header = line[1:].split()[0]
            parts = []
        else:
            parts.append(line.upper().replace("U", "T"))
    if header is not None:
        records[header] = "".join(parts)
    return records

def reverse_complement(seq):
    return seq.translate(_COMP)[::-1]

def gc_content(seq):
    s = seq.upper()
    return (s.count("G") + s.count("C")) / len(s) * 100

def tm_basic(seq):
    s = seq.upper()
    return 2 * (s.count("A") + s.count("T")) + 4 * (s.count("G") + s.count("C"))

def hairpin_score(seq):
    rc = reverse_complement(seq)
    best = 0
    for i in range(len(seq) - 4):
        for j in range(i + 5, len(seq) + 1):
            if seq[i:j] in rc:
                best = max(best, j - i)
    return best

def dimer_score(seq1, seq2):
    rc = reverse_complement(seq2)
    best = 0
    for i in range(len(seq1) - 4):
        for j in range(i + 5, len(seq1) + 1):
            if seq1[i:j] in rc:
                best = max(best, j - i)
    return best

def secondary_structure_score(target):
    rc = reverse_complement(target)
    return sum(1 for i in range(len(target) - 5) if target[i:i+6] in rc)

def passes_filters(gc, hp1, hp2, dim, struct):
    return (35 <= gc <= 65) and hp1 <= 6 and hp2 <= 6 and dim <= 6 and struct <= 4

# ── Core probe builder (single target) ────────────────────────────────────────
def build_probe_row(probe_num, target, arm_type, start=0, label=None):
    ini = INITIATORS[arm_type]
    P1, P2, S1, S2 = ini["P1"], ini["P2"], ini["S1"], ini["S2"]
    arm1   = target[:25]
    arm2   = target[-25:]
    probe1 = P1 + S1 + reverse_complement(arm1)
    probe2 = reverse_complement(arm2) + S2 + P2
    gc     = gc_content(target)
    hp1    = hairpin_score(probe1)
    hp2    = hairpin_score(probe2)
    dim    = dimer_score(probe1, probe2)
    struct = secondary_structure_score(target)
    tm1    = tm_basic(probe1)
    tm2    = tm_basic(probe2)
    if not passes_filters(gc, hp1, hp2, dim, struct):
        return None
    row = {
        "Probe_#":         probe_num,
        "Target_start":    start + 1,
        "Target_end":      start + 52,
        "Target_sequence": target,
        "Arm1_5to3":       arm1,
        "Arm2_5to3":       arm2,
        "Probe1_5to3":     probe1,
        "Probe2_5to3":     probe2,
        "Probe1_length":   len(probe1),
        "Probe2_length":   len(probe2),
        "GC_percent":      round(gc, 1),
        "Tm_P1_C":         round(tm1, 1),
        "Tm_P2_C":         round(tm2, 1),
        "Hairpin_P1":      hp1,
        "Hairpin_P2":      hp2,
        "Dimer_score":     dim,
        "Structure_score": struct,
        "Initiator_set":   arm_type,
    }
    if label is not None:
        row["Target_label"] = label
    return row

# ── Mode 1: single FASTA ───────────────────────────────────────────────────────
def generate_probes_from_fasta(seq, arm_type, step=52, target_len=52):
    rows = []
    for i in range(0, len(seq) - target_len + 1, step):
        target = seq[i: i + target_len]
        row = build_probe_row(len(rows) + 1, target, arm_type, start=i)
        if row:
            rows.append(row)
    return pd.DataFrame(rows)

def rank_and_select(df, max_probes=30):
    if df.empty:
        return df
    df = df.copy()
    df["_score"] = (
        abs(df["GC_percent"] - 50) * 0.5 +
        df["Hairpin_P1"] * 2 + df["Hairpin_P2"] * 2 +
        df["Dimer_score"] * 2 + df["Structure_score"] * 1
    )
    df = df.sort_values("_score").head(max_probes)
    df = df.sort_values("Target_start").reset_index(drop=True)
    df["Probe_#"] = range(1, len(df) + 1)
    return df.drop(columns=["_score"])

# ── Mode 2: direct 52 bp targets ──────────────────────────────────────────────
def parse_target_input(raw):
    targets = []
    lines = [l.strip() for l in raw.strip().splitlines() if l.strip()]
    i = 0
    label_pending = None
    while i < len(lines):
        line = lines[i]
        if line.startswith(">"):
            label_pending = line[1:].strip() or f"Target_{len(targets)+1:02d}"
            i += 1
            continue
        seq_line  = re.sub(r"^\d+[\.\)\:]\s*", "", line)
        seq_clean = re.sub(r"[^ACGTUacgtu]", "", seq_line).upper().replace("U", "T")
        if len(seq_clean) >= 52:
            label = label_pending or f"Target_{len(targets)+1:02d}"
            targets.append((label, seq_clean[:52]))
            label_pending = None
        i += 1
    return targets

def generate_probes_from_targets(targets, arm_type):
    rows, failed = [], []
    for idx, (label, target) in enumerate(targets):
        row = build_probe_row(idx + 1, target, arm_type, start=0, label=label)
        if row:
            rows.append(row)
        else:
            failed.append(label)
    df = pd.DataFrame(rows)
    if not df.empty:
        df["Probe_#"] = range(1, len(df) + 1)
    return df, failed

# ── Mode 3: multi-FASTA → common 52 bp regions → probes ───────────────────────
def find_common_regions(records: dict, k: int = 52) -> list:
    """
    Find all k-mers common to every sequence in records.
    Returns list of dicts with kmer, gc, positions per sequence,
    deduplicated to non-overlapping windows (step >= k).
    """
    names = list(records.keys())
    seqs  = list(records.values())

    # Build kmer sets
    kmer_sets = []
    for seq in seqs:
        kmer_sets.append(set(seq[i:i+k] for i in range(len(seq) - k + 1)))

    # Intersection across all sequences
    common = kmer_sets[0]
    for ks in kmer_sets[1:]:
        common &= ks

    if not common:
        return []

    # For each common kmer, record first position in each sequence and GC
    results = []
    for kmer in common:
        positions = {name: seq.index(kmer) + 1 for name, seq in zip(names, seqs)}
        gc = round(gc_content(kmer), 1)
        # Sort by position in first sequence
        results.append({
            "kmer":      kmer,
            "gc":        gc,
            "positions": positions,
            "pos_first": positions[names[0]]
        })

    results.sort(key=lambda x: x["pos_first"])

    # Deduplicate: keep non-overlapping (gap >= 0)
    deduped, last_pos = [], -9999
    for r in results:
        if r["pos_first"] - last_pos >= k:
            deduped.append(r)
            last_pos = r["pos_first"]

    return deduped

def common_regions_to_probe_df(common_regions, arm_type, seq_names):
    """Convert common region list to probe DataFrame, applying QC filters."""
    rows, failed = [], []
    for idx, region in enumerate(common_regions):
        kmer  = region["kmer"]
        label = f"Common_region_{idx+1:02d}"
        # Position note: pos in each sequence
        pos_str = " | ".join(f"{n}:{p}" for n, p in region["positions"].items())
        row = build_probe_row(idx + 1, kmer, arm_type, start=0, label=label)
        if row:
            row["GC_all_seqs"]    = region["gc"]
            row["Positions"]      = pos_str
            row["Same_position"]  = "Yes" if len(set(region["positions"].values())) == 1 else "No"
            rows.append(row)
        else:
            failed.append(label)
    df = pd.DataFrame(rows)
    if not df.empty:
        df["Probe_#"] = range(1, len(df) + 1)
    return df, failed

# ── Excel helpers ──────────────────────────────────────────────────────────────
def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _hfont(color=C_WHITE):
    return Font(name="Arial", size=11, bold=True, color=color)

def _cfont(bold=False):
    return Font(name="Arial", size=10, bold=bold, color=C_DARK)

def _center():
    return Alignment(horizontal="center", vertical="center")

def _left():
    return Alignment(horizontal="left", vertical="center")

def _rich_probe(probe_seq, initiator_part, spacer, spacer_after):
    nf = InlineFont(rFont="Arial", sz=20, color=C_DARK)
    sf = InlineFont(rFont="Arial", sz=20, color=C_SPACER)
    s_len = len(spacer)
    i_len = len(initiator_part)
    if not spacer_after:
        return CellRichText(
            TextBlock(nf, probe_seq[:i_len]),
            TextBlock(sf, probe_seq[i_len:i_len+s_len]),
            TextBlock(nf, probe_seq[i_len+s_len:]),
        )
    else:
        return CellRichText(
            TextBlock(nf, probe_seq[:len(probe_seq)-i_len-s_len]),
            TextBlock(sf, probe_seq[len(probe_seq)-i_len-s_len:len(probe_seq)-i_len]),
            TextBlock(nf, probe_seq[-i_len:]),
        )

# ── Sheet writers ──────────────────────────────────────────────────────────────
def write_info_sheet(ws, gene_name, arm_type, input_mode, extra_info, df):
    ini = INITIATORS[arm_type]
    ws.merge_cells("A1:D1")
    ws["A1"] = "HCR Probe Designer – Run Summary"
    ws["A1"].font = Font(name="Arial", size=13, bold=True, color=C_WHITE)
    ws["A1"].fill = _fill(C_TEAL)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    rows_info = [
        ("Input mode",      input_mode),
        ("Gene / Label",    gene_name),
        *extra_info,
        ("Initiator set",   arm_type),
        ("P1 initiator",    ini["P1"]),
        ("P2 initiator",    ini["P2"]),
        ("Spacer S1",       ini["S1"]),
        ("Spacer S2",       ini["S2"]),
        ("Probes output",   len(df)),
        ("Oligos to order", len(df) * 2),
        ("GC filter",       "35% – 65%"),
        ("Hairpin filter",  "≤ 6 nt"),
        ("Dimer filter",    "≤ 6 nt"),
        ("Structure filter","≤ 4 matches"),
    ]
    for i, (k, v) in enumerate(rows_info, start=2):
        ws.cell(row=i, column=1, value=k).font = Font(name="Arial", size=10, bold=True, color=C_DARK)
        ws.cell(row=i, column=1).fill = _fill(C_TEAL_LITE)
        ws.cell(row=i, column=2, value=v).font = Font(name="Arial", size=10, color=C_DARK)
        ws.cell(row=i, column=2).fill = _fill(C_WHITE)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 30

def write_common_regions_sheet(ws, common_regions, seq_names):
    """Extra sheet for Mode 3 only — shows all common regions before QC filter."""
    ws.merge_cells("A1:I1")
    ws["A1"] = f"All Common 52 bp Regions — {len(seq_names)} sequences"
    ws["A1"].font = Font(name="Arial", size=13, bold=True, color=C_WHITE)
    ws["A1"].fill = _fill(C_PURPLE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    HEADS  = ["#", "GC %", "GC Status", "Same Position?",
              *[f"Pos in {n}" for n in seq_names],
              "Common 52 bp Sequence"]
    # dynamic widths for position columns
    WIDTHS = [4, 7, 12, 14] + [14]*len(seq_names) + [56]
    for ci, (h, w) in enumerate(zip(HEADS, WIDTHS), start=1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font = _hfont(); cell.fill = _fill(C_DARK)
        cell.alignment = _center(); cell.border = _thin_border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 20

    for i, region in enumerate(common_regions, start=1):
        gc    = region["gc"]
        ok    = 35 <= gc <= 65
        bg    = C_GREEN_LT if ok else C_RED_LT
        same  = "Yes" if len(set(region["positions"].values())) == 1 else "No"
        row_vals = (
            [i, gc, "✔ PASS" if ok else "✘ High/Low GC", same]
            + [region["positions"][n] for n in seq_names]
            + [region["kmer"]]
        )
        for ci, v in enumerate(row_vals, start=1):
            cell = ws.cell(row=3+i, column=ci, value=v)
            cell.font = Font(name="Arial", size=10,
                             bold=(ci==3),
                             color=C_GREEN if (ok and ci==3) else C_RED if (not ok and ci==3) else C_DARK)
            cell.fill = _fill(bg)
            cell.border = _thin_border()
            cell.alignment = _left() if ci == len(row_vals) else _center()
    ws.freeze_panes = ws["A4"]

def write_design_sheet(ws, df, gene_name, arm_type, has_label_col=False, extra_cols=None):
    ini = INITIATORS[arm_type]
    n_extra = len(extra_cols) if extra_cols else 0
    total_cols = 18 + (1 if has_label_col else 0) + n_extra
    last_col = get_column_letter(total_cols)
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = f"HCR Probe Design Report  ·  {gene_name}  ·  Initiator: {arm_type}"
    ws["A1"].font = Font(name="Arial", size=13, bold=True, color=C_WHITE)
    ws["A1"].fill = _fill(C_TEAL)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A2:D2")
    ws["A2"] = f"Probes: {len(df)}   |   Oligos to order: {len(df)*2}"
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color=C_DARK)
    ws["A2"].fill = _fill(C_TEAL_LITE)
    ws["A2"].alignment = _left()
    ws.merge_cells(f"A4:{last_col}4")
    ws["A4"] = f"✔  PASSING PROBES  ({len(df)} probes ready for ordering)"
    ws["A4"].font = Font(name="Arial", size=11, bold=True, color=C_WHITE)
    ws["A4"].fill = _fill(C_GREEN)
    ws["A4"].alignment = _left()
    ws.row_dimensions[4].height = 22

    HEADERS    = ["#","Start","End","Target Sequence (52 nt)","Arm1 (25 nt)","Arm2 (25 nt)",
                  "Probe 1 (5'→3')","Probe 2 (5'→3')","P1 Len","P2 Len",
                  "GC %","Tm P1 (°C)","Tm P2 (°C)","Hairpin P1","Hairpin P2",
                  "Dimer","Structure","Initiator"]
    COL_WIDTHS = [5,7,7,34,16,16,42,42,7,7,8,10,10,10,10,8,10,10]
    KEYS       = ["Probe_#","Target_start","Target_end","Target_sequence",
                  "Arm1_5to3","Arm2_5to3","Probe1_5to3","Probe2_5to3",
                  "Probe1_length","Probe2_length","GC_percent","Tm_P1_C","Tm_P2_C",
                  "Hairpin_P1","Hairpin_P2","Dimer_score","Structure_score","Initiator_set"]

    if has_label_col:
        HEADERS    = ["Target Label"] + HEADERS
        COL_WIDTHS = [22] + COL_WIDTHS
        KEYS       = ["Target_label"] + KEYS

    if extra_cols:
        for col_key, col_head, col_w in extra_cols:
            HEADERS.append(col_head)
            COL_WIDTHS.append(col_w)
            KEYS.append(col_key)

    for ci, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=5, column=ci, value=h)
        cell.font = _hfont(); cell.fill = _fill(C_DARK)
        cell.alignment = _center(); cell.border = _thin_border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[5].height = 20

    label_offset = 1 if has_label_col else 0
    for r_off, (_, rec) in enumerate(df.iterrows()):
        row = 6 + r_off
        for ci, key in enumerate(KEYS, start=1):
            cell = ws.cell(row=row, column=ci)
            val  = rec.get(key, "")
            if key == "Probe1_5to3":
                cell.value = _rich_probe(val, ini["P1"], ini["S1"], spacer_after=False)
            elif key == "Probe2_5to3":
                cell.value = _rich_probe(val, ini["P2"], ini["S2"], spacer_after=True)
            else:
                cell.value = val
                cell.font  = _cfont(bold=(key in ("Probe_#", "Target_label")))
            cell.fill = _fill(C_GREEN_LT); cell.border = _thin_border()
            cell.alignment = _left() if ci > (3 + label_offset) else _center()
    ws.freeze_panes = ws["A6"]

def write_order_sheet(ws, df, gene_name):
    ws.merge_cells("A1:G1")
    ws["A1"] = f"Oligo Ordering Sheet  ·  {gene_name}  ·  HCR Probes"
    ws["A1"].font = Font(name="Arial", size=13, bold=True, color=C_WHITE)
    ws["A1"].fill = _fill(C_TEAL)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A2:G2")
    ws["A2"] = f"{len(df)} probe pairs  ·  {len(df)*2} oligos total  ·  All 5'→3', STD desalting"
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color=C_DARK)
    ws["A2"].fill = _fill(C_TEAL_LITE)
    ws["A2"].alignment = _left()
    HEADS  = ["Oligo Name","Sequence (5'→3')","Length (nt)","GC %","Tm (°C)","Purification","Notes"]
    WIDTHS = [32,52,12,8,10,14,40]
    for ci, (h, w) in enumerate(zip(HEADS, WIDTHS), start=1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font = _hfont(); cell.fill = _fill(C_DARK)
        cell.alignment = _center(); cell.border = _thin_border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 20

    r = 4
    for idx, rec in df.iterrows():
        pair_num   = idx + 1
        ini        = INITIATORS[rec["Initiator_set"]]
        label_base = rec.get("Target_label", gene_name)
        for probe_num, (seq_key, lbl) in enumerate(
                [("Probe1_5to3","P1"),("Probe2_5to3","P2")], start=1):
            seq  = rec[seq_key]
            name = f"{label_base}_{rec['Initiator_set']}_pair{pair_num:02d}_{lbl}"
            gc   = round(gc_content(seq), 1)
            tm   = round(tm_basic(seq), 1)
            note = f"Target: {rec['Target_sequence'][:20]}… | Arm{probe_num}"
            if rec.get("Positions"):
                note += f" | {rec['Positions']}"
            rich_seq = _rich_probe(seq,
                                   ini["P1"] if lbl=="P1" else ini["P2"],
                                   ini["S1"] if lbl=="P1" else ini["S2"],
                                   spacer_after=(lbl=="P2"))
            bg = C_GREEN_LT if probe_num == 1 else C_GREY
            for ci, v in enumerate([name, rich_seq, len(seq), gc, tm, "STD desalt", note], start=1):
                cell = ws.cell(row=r, column=ci, value=v)
                if ci != 2:
                    cell.font = _cfont(bold=(ci==1))
                cell.fill = _fill(bg); cell.border = _thin_border()
                cell.alignment = _left() if ci in (1,2,7) else _center()
            r += 1
        for ci in range(1, 8):
            ws.cell(row=r, column=ci).fill = _fill(C_WHITE)
        ws.row_dimensions[r].height = 4
        r += 1
    ws.freeze_panes = ws["A4"]

def build_excel(df, gene_name, arm_type, input_mode, extra_info,
                has_label_col=False, extra_cols=None,
                common_regions=None, seq_names=None):
    wb = Workbook()
    ws_info   = wb.active; ws_info.title = "Run_Info"
    if common_regions is not None:
        ws_common = wb.create_sheet("All_Common_Regions")
    ws_design = wb.create_sheet("Probe_Design")
    ws_order  = wb.create_sheet("Order_Sheet")
    write_info_sheet(ws_info, gene_name, arm_type, input_mode, extra_info, df)
    if common_regions is not None:
        write_common_regions_sheet(ws_common, common_regions, seq_names)
    write_design_sheet(ws_design, df, gene_name, arm_type, has_label_col, extra_cols)
    write_order_sheet(ws_order, df, gene_name)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

# ══════════════════════════════════════════════════════════════════════════════
#  STREAMLIT UI
# ══════════════════════════════════════════════════════════════════════════════

with st.sidebar:
    st.header("⚙️ Settings")
    arm_type = st.selectbox("Initiator Set", ["B1","B2","B3","B4","B5"], index=0)
    ini = INITIATORS[arm_type]
    st.caption(f"P1: `{ini['P1']}`")
    st.caption(f"P2: `{ini['P2']}`")
    st.caption(f"Spacers: S1=`{ini['S1']}`  S2=`{ini['S2']}`")
    st.divider()
    st.markdown("**QC Filters (fixed)**")
    st.caption("GC: 35–65%")
    st.caption("Hairpin: ≤ 6 nt")
    st.caption("Dimer: ≤ 6 nt")
    st.caption("Structure: ≤ 4 matches")
    st.caption("Max output (Mode 1): 30 best probes")

tab1, tab2, tab3 = st.tabs([
    "📄  Mode 1 — Single FASTA",
    "🎯  Mode 2 — Direct 52 bp Targets",
    "🔀  Mode 3 — Multi-FASTA Common Regions",
])

# ════════════════════════════════════════════════════════════
#  TAB 1 — SINGLE FASTA
# ════════════════════════════════════════════════════════════
with tab1:
    st.markdown("Paste a single gene sequence. The app tiles it into non-overlapping 52 bp windows, applies QC filters, and returns the top 30 probes.")

    fasta_input = st.text_area(
        "Paste FASTA sequence",
        height=200,
        placeholder=">gene_name\nATGGATGATGATATCGCCGCGCTCGTCGTCGAC...",
        key="fasta_input"
    )
    run_fasta = st.button("🚀 Design Probes from FASTA", type="primary",
                          use_container_width=True, key="run_fasta")

    if run_fasta:
        if not fasta_input.strip():
            st.error("Please paste a FASTA sequence above.")
        else:
            records = parse_fasta(fasta_input)
            if not records:
                st.error("Could not parse FASTA. Make sure it starts with a > header line.")
            else:
                gene_name, seq = next(iter(records.items()))
                seq = seq.upper().replace("U", "T")
                with st.spinner("Tiling and designing probes…"):
                    df = generate_probes_from_fasta(seq, arm_type, step=52)
                if df.empty:
                    st.warning("No probes passed QC filters. Try a different initiator set.")
                else:
                    df = rank_and_select(df, max_probes=30)
                    st.success(f"✅ {len(df)} probes designed — {len(df)*2} oligos ready to order")
                    c1, c2, c3, c4 = st.columns(4)
                    c1.metric("Gene",          gene_name)
                    c2.metric("Sequence (nt)", len(seq))
                    c3.metric("Probes",        len(df))
                    c4.metric("Oligos",        len(df)*2)
                    st.subheader("Probe Preview")
                    st.dataframe(df[["Probe_#","Target_start","Target_end","GC_percent",
                                     "Tm_P1_C","Tm_P2_C","Hairpin_P1","Hairpin_P2",
                                     "Dimer_score","Structure_score"]],
                                 use_container_width=True, hide_index=True)
                    extra = [("Sequence length (nt)", len(seq)),
                             ("Tiling step (nt)", 52), ("Max probes cap", 30)]
                    excel_bytes = build_excel(df, gene_name, arm_type,
                                             "Single FASTA — auto tiling", extra)
                    st.download_button(
                        "📥 Download Excel",
                        data=excel_bytes,
                        file_name=f"HCR_probes_{gene_name}_{arm_type}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )

# ════════════════════════════════════════════════════════════
#  TAB 2 — DIRECT 52 BP TARGETS
# ════════════════════════════════════════════════════════════
with tab2:
    st.markdown("Paste pre-selected 52 bp target regions. Add `>label` before each (FASTA-style) or just paste bare sequences.")
    st.info(
        "**With labels:**\n```\n>Region_06\nCGTCTTCCTCATGTACCTCTCCCAAACCCGGACCAAGTTGGCCGACGGACTG\n```\n"
        "**Plain:**\n```\nCGTCTTCCTCATGTACCTCTCCCAAACCCGGACCAAGTTGGCCGACGGACTG\n```",
        icon="💡"
    )
    gene_label = st.text_input("Gene / experiment label", value="MyGene", key="gene_label")
    target_input = st.text_area("Paste 52 bp target regions", height=250,
                                placeholder=">Region_06\nCGTCTTCCTCATGTACCTCTCCCAAACCCGGACCAAGTTGGCCGACGGACTG",
                                key="target_input")
    run_targets = st.button("🚀 Design Probes from Target Regions", type="primary",
                            use_container_width=True, key="run_targets")

    if run_targets:
        if not target_input.strip():
            st.error("Please paste at least one 52 bp target sequence above.")
        else:
            targets = parse_target_input(target_input)
            if not targets:
                st.error("No valid sequences found. Each sequence must be ≥ 52 nt.")
            else:
                with st.spinner(f"Designing probes for {len(targets)} target region(s)…"):
                    df, failed = generate_probes_from_targets(targets, arm_type)
                if failed:
                    st.warning(f"⚠️ {len(failed)} target(s) failed QC: " + ", ".join(failed))
                if df.empty:
                    st.error("All targets failed QC filters. Try a different initiator set.")
                else:
                    st.success(f"✅ {len(df)} probes designed — {len(df)*2} oligos ready to order")
                    c1, c2, c3, c4 = st.columns(4)
                    c1.metric("Label",           gene_label)
                    c2.metric("Targets supplied", len(targets))
                    c3.metric("Probes passed QC", len(df))
                    c4.metric("Oligos",           len(df)*2)
                    st.subheader("Probe Preview")
                    st.dataframe(df[["Probe_#","Target_label","Target_sequence","GC_percent",
                                     "Tm_P1_C","Tm_P2_C","Hairpin_P1","Hairpin_P2",
                                     "Dimer_score","Structure_score"]],
                                 use_container_width=True, hide_index=True)
                    extra = [("Targets supplied", len(targets)),
                             ("Targets passed QC", len(df)),
                             ("Targets failed QC", len(failed))]
                    excel_bytes = build_excel(df, gene_label, arm_type,
                                             "Direct 52 bp target regions", extra,
                                             has_label_col=True)
                    st.download_button(
                        "📥 Download Excel",
                        data=excel_bytes,
                        file_name=f"HCR_probes_{gene_label}_{arm_type}_targets.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )

# ════════════════════════════════════════════════════════════
#  TAB 3 — MULTI-FASTA COMMON REGIONS
# ════════════════════════════════════════════════════════════
with tab3:
    st.markdown(
        "Paste a **multi-FASTA file** containing 2 or more sequences. "
        "The app will first find all **non-overlapping 52 bp regions common to every sequence**, "
        "apply QC filters, then design HCR probes on the passing regions."
    )
    st.info(
        "Paste all sequences together — each starting with `>`:\n"
        "```\n>CD8A_isoform1\nATGGCCTCCTGGGTGACC...\n\n>CD8A_isoform2\nATGGCCTCCTGGGTGACC...\n```",
        icon="💡"
    )

    multi_label = st.text_input(
        "Gene / experiment label (used in oligo names)",
        value="MultiSeq", key="multi_label"
    )

    multi_fasta_input = st.text_area(
        "Paste multi-FASTA here (2+ sequences)",
        height=260,
        placeholder=">Isoform_1\nATGGCCTCCTGGGTGACC...\n\n>Isoform_2\nATGGCCTCCTGGGTGACC...",
        key="multi_fasta_input"
    )

    run_multi = st.button("🔍 Find Common Regions & Design Probes", type="primary",
                          use_container_width=True, key="run_multi")

    if run_multi:
        if not multi_fasta_input.strip():
            st.error("Please paste a multi-FASTA above.")
        else:
            records = parse_fasta(multi_fasta_input)
            if len(records) < 2:
                st.error("Need at least 2 sequences. Make sure each starts with a > header line.")
            else:
                seq_names = list(records.keys())
                st.info(f"Found **{len(records)} sequences**: {', '.join(f'`{n}`' for n in seq_names)}")

                # ── Step 1: find common regions ────────────────────────────────
                with st.spinner("Finding common 52 bp regions across all sequences…"):
                    common_regions = find_common_regions(records, k=52)

                if not common_regions:
                    st.error("No 52 bp regions found in common across all sequences.")
                else:
                    # Show common region summary table
                    st.subheader(f"Step 1 — Common Regions Found: {len(common_regions)}")
                    summary_rows = []
                    for i, r in enumerate(common_regions, 1):
                        row = {"#": i, "GC %": r["gc"],
                               "GC Status": "✔ PASS" if 35 <= r["gc"] <= 65 else "✘ Fail",
                               "Same Position?": "Yes" if len(set(r["positions"].values())) == 1 else "No",
                               "Sequence (52 bp)": r["kmer"]}
                        for n in seq_names:
                            row[f"Pos in {n}"] = r["positions"][n]
                        summary_rows.append(row)
                    summary_df = pd.DataFrame(summary_rows)
                    st.dataframe(summary_df, use_container_width=True, hide_index=True)

                    gc_pass = sum(1 for r in common_regions if 35 <= r["gc"] <= 65)
                    c1, c2, c3 = st.columns(3)
                    c1.metric("Total common regions", len(common_regions))
                    c2.metric("GC 35–65% (usable)",   gc_pass)
                    c3.metric("High/low GC (flagged)", len(common_regions) - gc_pass)

                    st.divider()

                    # ── Step 2: design probes on passing common regions ────────
                    with st.spinner("Designing HCR probes on passing common regions…"):
                        df, failed = common_regions_to_probe_df(
                            common_regions, arm_type, seq_names
                        )

                    if failed:
                        st.warning(f"⚠️ {len(failed)} region(s) failed QC filters: " + ", ".join(failed))

                    if df.empty:
                        st.error("No probes passed QC. Try a different initiator set.")
                    else:
                        st.subheader(f"Step 2 — Probes Designed: {len(df)}")
                        st.success(f"✅ {len(df)} probes — {len(df)*2} oligos ready to order")

                        c1, c2, c3, c4 = st.columns(4)
                        c1.metric("Sequences",         len(records))
                        c2.metric("Common regions",    len(common_regions))
                        c3.metric("Probes passed QC",  len(df))
                        c4.metric("Oligos",            len(df)*2)

                        preview_cols = ["Probe_#","Target_label","Target_sequence",
                                        "GC_percent","Tm_P1_C","Tm_P2_C",
                                        "Hairpin_P1","Hairpin_P2","Dimer_score",
                                        "Structure_score","Same_position","Positions"]
                        st.dataframe(df[[c for c in preview_cols if c in df.columns]],
                                     use_container_width=True, hide_index=True)

                        extra = [
                            ("Sequences in file",   len(records)),
                            ("Sequence names",       ", ".join(seq_names)),
                            ("Common regions found", len(common_regions)),
                            ("Regions passed QC",    len(df)),
                            ("Regions failed QC",    len(failed)),
                        ]
                        extra_cols = [
                            ("Same_position", "Same Position?", 16),
                            ("Positions",     "Positions in each sequence", 45),
                        ]
                        excel_bytes = build_excel(
                            df, multi_label, arm_type,
                            "Multi-FASTA — common 52 bp regions", extra,
                            has_label_col=True,
                            extra_cols=extra_cols,
                            common_regions=common_regions,
                            seq_names=seq_names,
                        )
                        st.download_button(
                            "📥 Download Excel (All_Common_Regions + Probe_Design + Order_Sheet)",
                            data=excel_bytes,
                            file_name=f"HCR_probes_{multi_label}_{arm_type}_common.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                        )
